import asyncio, aiohttp, json, re, csv
from datetime import datetime
from typing import Any, Dict, Tuple, Optional, List
from urllib.parse import urlparse
from pathlib import Path

# ====== Telegram ayarları ======
TELEGRAM_BOT_TOKEN = "7895901821:AAEJs3mmWxiWrRyVKcRiAxMN2Rn4IpiyV0o"
TELEGRAM_CHAT_ID   = "-4678220102"

# ====== Script klasörü ======
BASE_DIR = Path(__file__).resolve().parent

# ----- Deposit dosyaları -----
DEPOSIT_API_URL_FILE   = BASE_DIR / "api_url.txt"              # https://.../admin-api/depositCrypto
DEPOSIT_PAGE_URL_FILE  = BASE_DIR / "page_url.txt"             # https://.../depositCrypto
DEPOSIT_COOKIE_FILE    = BASE_DIR / "cookie.txt"               # ortak
DEPOSIT_PAYLOAD_FILE   = BASE_DIR / "payload.json"             # opsiyonel
DEPOSIT_STATE_FILE     = BASE_DIR / "state.json"

# ----- Withdraw dosyaları -----
WITHDRAW_API_URL_FILE  = BASE_DIR / "api_url_withdraw.txt"     # https://.../admin-api/withdrawCrypto
WITHDRAW_PAGE_URL_FILE = BASE_DIR / "page_url_withdraw.txt"    # https://.../withdrawCrypto
WITHDRAW_PAYLOAD_FILE  = BASE_DIR / "payload_withdraw.json"    # opsiyonel
WITHDRAW_STATE_FILE    = BASE_DIR / "state_withdraw.json"

# ----- CountryCode için ORTAK -----
CC_PAGE_URL_FILE      = BASE_DIR / "cc_page_url.txt"           # baz: https://.../userDetail?id=
CC_API_URL_FILE       = BASE_DIR / "cc_api_url.txt"            # https://.../admin-api/get_user_info

# ----- Ülke isim eşleme dosyaları (CSV / XLSX) -----
COUNTRY_MAP_CSV       = BASE_DIR / "country_map.csv"           # Code,Country (opsiyonel)
COUNTRY_MAP_XLSX      = BASE_DIR / "country_map.xlsx"          # COUNTRY, COUNTRY CODE (tercih edilen)

# ====== Ayarlar ======
POLL_SECONDS   = 5
SEED_ON_START  = True
CC_TIMEOUT_SEC = 1.5     # countryCode getirirken max bekleme sn

# ----------------------------------------------------------
def read_text_or_empty(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8-sig").strip()
    except FileNotFoundError:
        return ""

def normalize_cookie_value(text: str) -> str:
    if not text:
        return ""
    t = text.strip()
    if t.lower().startswith("cookie:"):
        t = t.split(":", 1)[1].strip()
    return " ".join(s.strip() for s in t.splitlines() if s.strip())

def cookie_get(cookie_value: str, name: str) -> Optional[str]:
    if not cookie_value:
        return None
    for part in cookie_value.split(";"):
        if "=" not in part:
            continue
        k, v = part.split("=", 1)
        if k.strip().lower() == name.lower():
            return v.strip()
    return None

def parse_origin(url: str) -> str:
    try:
        u = urlparse(url)
        if not u.scheme or not u.netloc:
            return ""
        host = u.hostname or ""
        port = ""
        if u.port and not ((u.scheme == "https" and u.port == 443) or (u.scheme == "http" and u.port == 80)):
            port = f":{u.port}"
        return f"{u.scheme}://{host}{port}"
    except Exception:
        return ""

def load_json_or(path: Path, default: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return default

def save_json(path: Path, obj: Any) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)

def fmt_tr(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y %H.%M.%S")

class FileWatcher:
    def __init__(self, path: Path, normalize_fn=None):
        self.path = Path(path)
        self.norm = normalize_fn
        self.mtime = 0.0
        self.value: Optional[str] = None

    def load_if_changed(self) -> Tuple[bool, Optional[str]]:
        try:
            st = self.path.stat()
            if self.mtime != st.st_mtime:
                raw = read_text_or_empty(self.path)
                self.value = self.norm(raw) if self.norm else raw
                self.mtime = st.st_mtime
                return True, self.value
        except FileNotFoundError:
            pass
        return False, self.value

# ---------- Country map (CSV veya XLSX) ----------
class CountryMap:
    """
    - XLSX varsa onu okur (COUNTRY, COUNTRY CODE sütunları).
    - Yoksa CSV okur (Code,Country / Country Code,Country vs. esnek başlık).
    - Kod hücresinde '1-809, 1-829, 1-849' gibi ifadeler varsa içindeki TÜM sayı parçalarını ayrı ayrı map'ler.
    """
    def __init__(self, csv_path: Path, xlsx_path: Path):
        self.csv_path  = Path(csv_path)
        self.xlsx_path = Path(xlsx_path)
        self.use_xlsx  = self.xlsx_path.exists()
        self.mtime = 0.0
        self.map: Dict[str, List[str]] = {}

    # --- yardımcılar ---
    @staticmethod
    def _numeric_pieces(cell: str) -> List[str]:
        """Hücre içindeki TÜM sayı parçalarını çıkar (örn '1-829, 1-849' -> ['1','829','1','849'])."""
        if not cell:
            return []
        return re.findall(r"\d+", str(cell))

    @staticmethod
    def _append(mapobj: Dict[str, List[str]], code: str, country: str):
        if not code or not country:
            return
        mapobj.setdefault(code, [])
        if country not in mapobj[code]:
            mapobj[code].append(country)

    # --- CSV ---
    def _load_csv(self) -> Dict[str, List[str]]:
        res: Dict[str, List[str]] = {}
        with self.csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return res
        headers = [h.strip().lower() for h in rows[0]]
        # sütun indeksleri
        try:
            code_idx = headers.index("country code")
        except ValueError:
            try:
                code_idx = headers.index("code")
            except ValueError:
                code_idx = 0
        try:
            country_idx = headers.index("country")
        except ValueError:
            country_idx = 1 if len(headers) > 1 else 0

        for row in rows[1:]:
            if len(row) <= max(code_idx, country_idx):
                continue
            country = (row[country_idx] or "").strip()
            for piece in self._numeric_pieces(row[code_idx]):
                self._append(res, piece, country)
        return res

    # --- XLSX ---
    def _load_xlsx(self) -> Dict[str, List[str]]:
        res: Dict[str, List[str]] = {}
        try:
            import openpyxl  # pip install openpyxl
        except Exception as e:
            print(f"[CountryMap] openpyxl yok veya yüklenemedi: {e}")
            return res

        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True, read_only=True)
        ws = wb.active

        # Başlık satırı
        headers = []
        for cell in ws[1]:
            headers.append((cell.value or "").strip().lower())
        # Sütun indeksleri
        def idx_of(names: List[str], fallback: int) -> int:
            for i, h in enumerate(headers):
                if h in names:
                    return i
            return fallback

        code_col = idx_of(["country code", "country_code", "code", "kod"], 1)
        country_col = idx_of(["country", "ülke", "name"], 0)

        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = ["" if v is None else str(v) for v in row]
            if len(vals) <= max(code_col, country_col):
                continue
            country = vals[country_col].strip()
            code_cell = vals[code_col]
            for piece in self._numeric_pieces(code_cell):
                self._append(res, piece, country)
        return res

    def reload_if_changed(self) -> bool:
        # hangi dosya?
        self.use_xlsx = self.xlsx_path.exists()
        target = self.xlsx_path if self.use_xlsx else self.csv_path
        try:
            st = target.stat()
            if st.st_mtime == self.mtime:
                return False
            self.mtime = st.st_mtime
        except FileNotFoundError:
            if self.map:
                self.map = {}
            return False

        try:
            new_map = self._load_xlsx() if self.use_xlsx else self._load_csv()
            self.map = new_map
            kind = "XLSX" if self.use_xlsx else "CSV"
            print(f"[CountryMap] {kind} yüklendi; toplam kod={len(self.map)}")
            return True
        except Exception as e:
            print(f"[CountryMap] yükleme hatası: {e}")
            self.map = {}
            return True

    def countries_for_cc(self, cc_str: Optional[str]) -> Optional[str]:
        """cc_str: '+792+90' -> '90' al, map'te ara, listeyi ' / ' ile birleştir."""
        if not cc_str:
            return None
        digits = re.findall(r"\d+", str(cc_str))
        if not digits:
            return None
        code = digits[-1]
        lst = self.map.get(code)
        if not lst:
            return None
        return " / ".join(lst)

# ---------- Telegram ----------
class Telegram:
    def __init__(self, session: aiohttp.ClientSession, token: str, chat_id: str):
        self.session = session
        self.token = token
        self.chat_id = chat_id

    async def send(self, text: str) -> None:
        print(f"[TG] {text}")
        if not self.token or not self.chat_id:
            return
        url = f"https://api.telegram.org/bot{self.token}/sendMessage"
        data = {"chat_id": self.chat_id, "text": text, "disable_web_page_preview": True}
        try:
            async with self.session.post(url, json=data, timeout=aiohttp.ClientTimeout(total=15)) as r:
                await r.text()
        except Exception as e:
            print(f"[WARN] Telegram gönderilemedi: {e}")

# ---------- Yardımcılar ----------
def build_user_detail_url(base_or_full: str, uid: str) -> Optional[str]:
    if not base_or_full or not uid:
        return None
    s = base_or_full.strip()
    if re.search(r"id=$", s):
        return s + uid
    if re.search(r"id=\d+$", s):
        return re.sub(r"id=\d+$", f"id={uid}", s)
    if "?" in s:
        joiner = "&" if not s.endswith("&") else ""
        return f"{s}{joiner}id={uid}"
    else:
        joiner = "?" if not s.endswith("?") else ""
        return f"{s}{joiner}id={uid}"

class BaseWatcher:
    def __init__(
        self,
        name: str,
        api_url_file: Path,
        page_url_file: Path,
        cookie_file: Path,
        payload_file: Path,
        state_file: Path,
        list_keys: List[str],
        header_emoji: str,
        header_title: str,
        wanted_fields_fn,
        cc_page_fw: FileWatcher,
        cc_api_fw: FileWatcher,
        country_map: CountryMap,
    ):
        self.name = name
        self.api_fw     = FileWatcher(api_url_file,   normalize_fn=lambda s: s.splitlines()[0].strip() if s else "")
        self.page_fw    = FileWatcher(page_url_file,  normalize_fn=lambda s: s.splitlines()[0].strip() if s else "")
        self.cookie_fw  = FileWatcher(cookie_file,    normalize_fn=normalize_cookie_value)
        self.payload_fw = FileWatcher(payload_file,   normalize_fn=lambda s: s)

        self.api_url  = ""
        self.page_url = ""
        self.cookie   = ""
        self.payload  = {}

        st = load_json_or(state_file, {"seen_ids": [], "seeded": False})
        self.seen_ids_file = state_file
        self.seen_ids = set(st.get("seen_ids", []))
        self.seeded   = bool(st.get("seeded", False))
        self.logged_out = False

        self.list_keys = list_keys
        self.header_emoji = header_emoji
        self.header_title = header_title
        self.make_message = wanted_fields_fn

        self.cc_page_fw = cc_page_fw
        self.cc_api_fw  = cc_api_fw
        self.country_map = country_map

        self.session: Optional[aiohttp.ClientSession] = None
        self.tg: Optional[Telegram] = None

    async def _get_html(self, url: str, origin: str, cookie: str, admin_token: Optional[str]) -> Tuple[int, str]:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Safari",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Origin": origin,
            "Cookie": cookie,
            "Cache-Control": "no-cache",
        }
        if admin_token:
            headers["Admin-Token"] = admin_token
        try:
            async with self.session.get(url, headers=headers) as r:
                txt = await r.text()
                return r.status, txt[:2000]
        except Exception as e:
            return 0, f"[HTML hata] {e}"

    async def _post_api(self, url: str, origin: str, page_url: str, cookie: str, payload: Dict[str, Any]) -> Tuple[int, Dict[str, Any] | str]:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Safari",
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json",
            "Accept-Language": "en-US,en;q=0.9",
            "Origin": origin,
            "Referer": page_url,
            "X-Requested-With": "XMLHttpRequest",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "Cookie": cookie,
            "Admin-language": "en_US",
        }
        admin_token = cookie_get(cookie, "admin-token")
        if admin_token:
            headers["Admin-Token"] = admin_token
        csrf = cookie_get(cookie, "csrfToken")
        if csrf:
            headers["X-CSRF-Token"] = csrf
        token_swap = cookie_get(cookie, "admin-token-swap")
        if token_swap:
            headers["Admin-Token-Swap"] = token_swap

        try:
            async with self.session.post(url, json=payload, headers=headers) as r:
                txt = await r.text()
                try:
                    return r.status, json.loads(txt)
                except json.JSONDecodeError:
                    return r.status, txt[:2000]
        except Exception as e:
            return 0, f"[API hata] {e}"

    async def _fetch_country_code_for_uid(self, uid: str) -> Optional[str]:
        if not uid:
            return None
        cc_page_base = self.cc_page_fw.value or ""
        cc_api_url   = self.cc_api_fw.value or ""
        if not cc_page_base or not cc_api_url:
            return None
        user_detail_url = build_user_detail_url(cc_page_base, uid)
        if not user_detail_url:
            return None
        origin = parse_origin(cc_api_url) or parse_origin(user_detail_url)
        if not origin:
            return None
        admin_token = cookie_get(self.cookie, "admin-token")
        try:
            _status_html, _ = await self._get_html(user_detail_url, origin, self.cookie, admin_token)
        except Exception as e:
            print(f"[CC][{self.name}] userDetail GET hata: {e}")
        await asyncio.sleep(0.25)
        payload = {"userId": uid}
        status, body = await self._post_api(
            url=cc_api_url,
            origin=origin,
            page_url=user_detail_url,
            cookie=self.cookie,
            payload=payload,
        )
        if isinstance(body, dict) and str(body.get("code")) == "0":
            data = body.get("data") or {}
            user = data.get("user") or {}
            cc = user.get("countryCode") or data.get("countryCode") or user.get("mobileNumberCountryCode")
            if cc not in (None, "", "null"):
                return str(cc)
        print(f"[CC][{self.name}] uid={uid} cc alınamadı; http={status}, body={str(body)[:300]}")
        return None

    def _find_list(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        for key in self.list_keys:
            arr = data.get(key)
            if isinstance(arr, list):
                return arr
        return []

    async def _handle_rows(self, body: Dict[str, Any]) -> None:
        data = body.get("data") or {}
        items = self._find_list(data)
        if not isinstance(items, list):
            await self.tg.send(f"⚠️ {self.name}: Beklenen liste bulunamadı ({' / '.join(self.list_keys)}).")
            return

        if SEED_ON_START and not self.seeded:
            for d in items:
                try:
                    self.seen_ids.add(int(d.get("id")))
                except Exception:
                    continue
            self.seeded = True
            save_json(self.seen_ids_file, {"seen_ids": sorted(self.seen_ids), "seeded": True})
            await self.tg.send(f"ℹ️ {self.name}: İlk yükleme — mevcut kayıtlar baz alındı (bildirim yok).")
            return

        new_items = []
        for d in items:
            try:
                did = int(d.get("id"))
            except Exception:
                continue
            if did not in self.seen_ids:
                new_items.append(d)
                self.seen_ids.add(did)

        if not new_items:
            return

        new_items.sort(key=lambda x: x.get("createdAt", 0))
        for d in new_items:
            uid = str(d.get("uid", "") or "")
            cc = "—"
            if uid:
                try:
                    cc = await asyncio.wait_for(self._fetch_country_code_for_uid(uid), timeout=CC_TIMEOUT_SEC)
                    if cc is None or cc == "" or str(cc).lower() == "null":
                        cc = "—"
                except asyncio.TimeoutError:
                    print(f"[CC][{self.name}] uid={uid} timeout ({CC_TIMEOUT_SEC}s). Mesaj cc'siz gönderilecek.")
                    cc = "—"
                except Exception as e:
                    print(f"[CC][{self.name}] uid={uid} hata: {e}")
                    cc = "—"

            # Ülke isimleri (xlsx/csv)
            self.country_map.reload_if_changed()
            country_names = self.country_map.countries_for_cc(cc if cc != "—" else None)
            country_line = f"Country: {country_names if country_names else '—'}"

            msg = self.make_message(d, self.header_emoji, self.header_title)
            parts = msg.splitlines()
            if parts:
                parts.insert(1, country_line)
                msg = "\n".join(parts)
            else:
                msg = f"{country_line}\n{msg}"
            msg = f"{msg}\ncountryCode: {cc}"
            await self.tg.send(msg)

        if len(self.seen_ids) > 5000:
            self.seen_ids = set(list(self.seen_ids)[-3000:])
        save_json(self.seen_ids_file, {"seen_ids": sorted(self.seen_ids), "seeded": self.seeded})

    async def run(self, session: aiohttp.ClientSession, tg: Telegram):
        self.session = session
        self.tg = tg

        for fw in (self.api_fw, self.page_fw, self.cookie_fw, self.payload_fw, self.cc_page_fw, self.cc_api_fw):
            fw.load_if_changed()

        if self.payload_fw.value:
            try:
                self.payload = json.loads(self.payload_fw.value)
            except Exception:
                self.payload = {}
                await self.tg.send(f"⚠️ {self.name}: payload JSON değil, boş gövde ile devam.")
        else:
            self.payload = {}

        self.api_url  = self.api_fw.value or ""
        self.page_url = self.page_fw.value or ""
        self.cookie   = self.cookie_fw.value or ""

        def info(p: Path) -> str:
            return f"{p}  (exists={p.exists()} size={p.stat().st_size if p.exists() else 0})"
        await self.tg.send(
            f"📂 {self.name} dosyaları:\n"
            f"- api: {info(self.api_fw.path)}\n"
            f"- page: {info(self.page_fw.path)}\n"
            f"- cookie: {info(self.cookie_fw.path)}\n"
            f"- payload: {info(self.payload_fw.path)}\n"
            f"- cc_page: {info(self.cc_page_fw.path)}\n"
            f"- cc_api: {info(self.cc_api_fw.path)}\n"
            f"- country_map.csv: {info(COUNTRY_MAP_CSV)}\n"
            f"- country_map.xlsx: {info(COUNTRY_MAP_XLSX)}"
        )

        while True:
            for name, fw in (("API URL", self.api_fw), ("PAGE URL", self.page_fw),
                             ("COOKIE", self.cookie_fw), ("PAYLOAD", self.payload_fw),
                             ("CC PAGE", self.cc_page_fw), ("CC API", self.cc_api_fw)):
                changed, val = fw.load_if_changed()
                if changed:
                    if name == "API URL":
                        self.api_url = val or ""
                    elif name == "PAGE URL":
                        self.page_url = val or ""
                    elif name == "COOKIE":
                        self.cookie = val or ""
                    elif name == "PAYLOAD":
                        try:
                            self.payload = json.loads(val) if val else {}
                        except Exception:
                            self.payload = {}
                            await self.tg.send(f"⚠️ {self.name}: payload JSON değil; boş gövde ile devam.")
                    await self.tg.send(f"ℹ️ {self.name}: {name} güncellendi.")

            if not self.api_url or not self.page_url or not self.cookie:
                await asyncio.sleep(POLL_SECONDS)
                continue

            origin = parse_origin(self.api_url or self.page_url)

            admin_token = cookie_get(self.cookie, "admin-token")
            h_status, _ = await self._get_html(self.page_url, origin, self.cookie, admin_token)
            print(f"[HTML][{self.name}] {self.page_url} -> {h_status}")

            a_status, a_body = await self._post_api(self.api_url, origin, self.page_url, self.cookie, self.payload)
            print(f"[API ][{self.name}] {self.api_url} -> {a_status}")

            if isinstance(a_body, dict):
                code = str(a_body.get("code"))
                if code == "0":
                    if self.logged_out:
                        await self.tg.send(f"✅ {self.name}: Oturum geri geldi (code 0).")
                        self.logged_out = False
                    await self._handle_rows(a_body)
                elif code == "10004":
                    if not self.logged_out:
                        await self.tg.send(f"🔴 {self.name}: User is not logged in (10004). Cookie/URL/headers expired.")
                        self.logged_out = True
                else:
                    await self.tg.send(f"⚠️ {self.name}: API code={code}, msg={a_body.get('msg')}")
            else:
                low = str(a_body).lower()
                if not self.logged_out and (a_status in (401,403) or "not logged in" in low):
                    await self.tg.send(f"🔴 {self.name}: Oturum geçersiz (HTTP {a_status}). Cookie/URL güncelleyin.")
                    self.logged_out = True

            await asyncio.sleep(POLL_SECONDS)

# ----- Deposit mesajı -----
def make_deposit_message(d: Dict[str, Any], emoji: str, title: str) -> str:
    try:
        ms = int(d.get("createdAt", 0))
    except Exception:
        ms = 0
    dt = datetime.fromtimestamp(ms/1000) if ms else datetime.now()
    tstr = fmt_tr(dt)

    symbol = str(d.get("symbol", ""))
    amount = str(d.get("amount", ""))
    txid = str(d.get("txid", ""))
    address_to = str(d.get("addressTo", ""))
    usdt_amount = str(d.get("usdtAmount", ""))
    uid = str(d.get("uid", ""))
    status_desc = str(d.get("statusDesc", d.get("walletStatus", "")))

    lines = [
        f"{emoji} {title}",
        f"symbol: {symbol}",
        f"amount: {amount}",
        f"txid: {txid}",
        f"addressTo: {address_to}",
        f"usdtAmount: {usdt_amount}",
        f"uid: {uid}",
        f"statusDesc: {status_desc}",
        f"time: {tstr}",
    ]
    coin_tx_url = str(d.get("coinTxUrl", "") or "")
    if coin_tx_url and txid:
        lines.append(f"txLink: {coin_tx_url}{txid}")
    return "\n".join(lines)

# ----- Withdraw mesajı -----
def make_withdraw_message(d: Dict[str, Any], emoji: str, title: str) -> str:
    try:
        ms = int(d.get("createdAt", 0))
    except Exception:
        ms = 0
    dt = datetime.fromtimestamp(ms/1000) if ms else datetime.now()
    tstr = fmt_tr(dt)

    symbol = str(d.get("symbol", ""))
    usdt_amount = str(d.get("usdtAmount", ""))
    txid = str(d.get("txid", ""))
    address_to = str(d.get("addressTo", ""))
    uid = str(d.get("uid", ""))
    amount = str(d.get("amount", ""))
    status_desc = str(d.get("statusDesc", ""))

    lines = [
        f"{emoji} {title}",
        f"symbol: {symbol}",
        f"usdtAmount: {usdt_amount}",
        f"txid: {txid}",
        f"addressTo: {address_to}",
        f"uid: {uid}",
        f"amount: {amount}",
        f"statusDesc: {status_desc}",
        f"time: {tstr}",
    ]
    coin_tx_url = str(d.get("coinTxUrl", "") or "")
    if coin_tx_url and txid:
        lines.append(f"txLink: {coin_tx_url}{txid}")
    return "\n".join(lines)

async def main():
    async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=25)) as session:
        tg = Telegram(session, TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID)
        await tg.send("🚀 Watcher’lar başlıyor (deposit + withdraw).")

        cc_page_fw = FileWatcher(CC_PAGE_URL_FILE, normalize_fn=lambda s: s.splitlines()[0].strip() if s else "")
        cc_api_fw  = FileWatcher(CC_API_URL_FILE,  normalize_fn=lambda s: s.splitlines()[0].strip() if s else "")

        country_map = CountryMap(COUNTRY_MAP_CSV, COUNTRY_MAP_XLSX)
        country_map.reload_if_changed()

        deposit = BaseWatcher(
            name="Deposit",
            api_url_file=DEPOSIT_API_URL_FILE,
            page_url_file=DEPOSIT_PAGE_URL_FILE,
            cookie_file=DEPOSIT_COOKIE_FILE,
            payload_file=DEPOSIT_PAYLOAD_FILE,
            state_file=DEPOSIT_STATE_FILE,
            list_keys=["depositCryptoMapList", "depositList"],
            header_emoji="🟢",
            header_title="Yeni Deposit",
            wanted_fields_fn=make_deposit_message,
            cc_page_fw=cc_page_fw,
            cc_api_fw=cc_api_fw,
            country_map=country_map,
        )

        withdraw = BaseWatcher(
            name="Withdraw",
            api_url_file=WITHDRAW_API_URL_FILE,
            page_url_file=WITHDRAW_PAGE_URL_FILE,
            cookie_file=DEPOSIT_COOKIE_FILE,
            payload_file=WITHDRAW_PAYLOAD_FILE,
            state_file=WITHDRAW_STATE_FILE,
            list_keys=["withdrawCryptoMapList", "withdrawList"],
            header_emoji="🔴",
            header_title="Yeni Withdraw",
            wanted_fields_fn=make_withdraw_message,
            cc_page_fw=cc_page_fw,
            cc_api_fw=cc_api_fw,
            country_map=country_map,
        )

        await asyncio.gather(
            deposit.run(session, tg),
            withdraw.run(session, tg),
        )

if __name__ == "__main__":
    asyncio.run(main())
